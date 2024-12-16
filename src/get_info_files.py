import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tabulate import tabulate
import datetime
import os
import sys

# Obtener la ruta absoluta del directorio actual
current_dir = os.path.dirname(os.path.abspath(__file__))

# Agregar la ruta al directorio src a la ruta de búsqueda de módulos
sys.path.append(os.path.join(current_dir, 'src'))


def Getinfo(archivos):

    archivos
    
# Obtener la fecha actual en formato mes-día
    date = datetime.datetime.now().strftime("%m-%d")
    
    # Cargar archivos de Excel
    coois = openpyxl.load_workbook(archivos[0])
    pct = openpyxl.load_workbook(archivos[1])
    master = openpyxl.load_workbook(archivos[2])
    ship = openpyxl.load_workbook(archivos[3])
    wop = openpyxl.load_workbook(archivos[4])
    prev_wop = openpyxl.load_workbook(archivos[5])

    # Cargar hojas
    coois_sheet = coois.active
    pct_sheet = pct.active
    master_sheet = master.active
    ship_sheet = ship.active
    wop_sheet_one = wop['Hoja1']
    wop_sheet_two = wop['Hoja2']
    prewop_sheet = prev_wop['Hoja1']

    # Obtener encabezado de cada archivo
    coois_headers = [cell.value for cell in coois_sheet[1]]
    master_headers = [cell.value for cell in master_sheet[1]]
    pct_headers = [cell.value for cell in pct_sheet[1]]
    ship_headers = [cell.value for cell in ship_sheet[1]]
    wop_headers_one = [cell.value for cell in wop_sheet_one[1]]
    wop_headers_two = [cell.value for cell in wop_sheet_two[1]]
    prewop_headers = [cell.value for cell in prewop_sheet[1]]


    # Obtener índices de las columnas de interés
    coois_wo_index = coois_headers.index("WO")  # Índice de "WO" en coois.xlsx
    coll_index = coois_headers.index("coll.do")
    prio_column_index = coois_headers.index("prio")
    master_wo_index = master_headers.index("WO")  # Índice de "WO" en master.xlsx
    root_column_index = pct_headers.index("RootCause")
    pct_wo_index = pct_headers.index("WO")  # Índice de "WO" en pct.xlsx
    ship_status_index = ship_headers.index("GENERAL STATUS")
    ship_wo_index = ship_headers.index("WO")
    wop_status_index = wop_headers_one.index("STATUS")
    wop_wo_index_one = wop_headers_one.index("WO")
    wop_wip_index = wop_headers_two.index("WIP")
    wop_wo_index_two = wop_headers_two.index("WO")
    prewop_status_index = prewop_headers.index("STATUS")
    prewop_wo_index = prewop_headers.index("WO")


    # Obtener valores únicos de la columna "WO" en master.xlsx como conjunto
    master_wo_set = {
        row[master_wo_index].value for row in master_sheet.iter_rows(min_row=2, max_row=master_sheet.max_row)
        if row[master_wo_index].value is not None
    }

    # Crear un diccionario para combinar datos de coois y pct por WO
    combined_data = {}

    # Filtrar y combinar datos de coois.xlsx
    for row in coois_sheet.iter_rows(min_row=2, max_row=coois_sheet.max_row):
        wo = row[coois_wo_index].value
        prio = row[prio_column_index].value
        if wo in master_wo_set:
            combined_data[wo] = [
                wo,  # WO
                row[coll_index].value,  # coll.do
                prio,  # prio
                None  # RootCause (se rellenará más tarde)
            ]


        
    # Filtrar datos de pct.xlsx y añadir RootCause al diccionario
    for row in pct_sheet.iter_rows(min_row=2, max_row=pct_sheet.max_row):
        wo = row[pct_wo_index].value
        if wo in master_wo_set and wo in combined_data:
            combined_data[wo][3] = row[root_column_index].value  # Añadir RootCause

    for row in ship_sheet.iter_rows(min_row=2, max_row=ship_sheet.max_row):
        wo = row[ship_wo_index].value
        if wo in master_wo_set and wo in combined_data:
            # Añadir nueva columna
            combined_data[wo].append(row[ship_status_index].value)

    for row in wop_sheet_one.iter_rows(min_row=2, max_row=wop_sheet_one.max_row):
        wo = row[wop_wo_index_one].value
        if wo in master_wo_set:  # Asegurar que WO esté en el set filtrado
            if wo not in combined_data:
                # Inicializar estructura
                combined_data[wo] = [wo, None, None, None, None]

            # Añadir 'Status' de Hoja1
            combined_data[wo].append(row[wop_status_index].value)

    # Filtrar y combinar datos de Hoja2
    for row in wop_sheet_two.iter_rows(min_row=2, max_row=wop_sheet_two.max_row):
        wo = row[wop_wo_index_two].value
        if wo in master_wo_set and wo in combined_data:  # Verificar que WO ya está en el dict
            # Añadir 'Comments' de Hoja2
            combined_data[wo].append(row[wop_wip_index].value)

    for row in prewop_sheet.iter_rows(min_row=2, max_row=prewop_sheet.max_row):
        wo = row[prewop_wo_index].value
        if wo in master_wo_set and wo in combined_data:
            # Añadir nueva columna
            combined_data[wo].append(row[prewop_status_index].value)



    for wo, data in combined_data.items():
        status = data[5]  # STATUS de Hoja1 (columna 6)
        prev_status = data[7]  # PREV STATUS de Hoja2 (columna 7)
        
        # Comparación de los dos valores
        if status == prev_status:
            data.append("TRUE")
        else:
            data.append("FALSE")

            

    # Convertir el diccionario a una lista para impresión/tabulación
    final_data = list(combined_data.values())

    # Imprimir los datos combinados
    headers = ["WO", "coll.do", "prio", "RootCause", "GENERAL STATUS", "STATUS", "WIP", "PREV STATUS", "IS STATUS CHANGED"]
    print("\nDatos combinados (filtrados por WO):")
    print(tabulate(final_data, headers=headers,
        tablefmt="fancy_grid", colalign=("center",) * len(headers)))

    # Crear nuevo archivo PCT MASTER TRACKER
    pct_master_tracker = openpyxl.Workbook()
    pct_master_tracker_sheet = pct_master_tracker.active
    pct_master_tracker_sheet.title = "Filtered Data"

    # Agregar encabezados al archivo
    for col, header in enumerate(headers, start=1):
        pct_master_tracker_sheet.cell(row=1, column=col).value = header

    # Insertar datos combinados en el nuevo archivo
    for i, data in enumerate(final_data, start=2):
        for j, value in enumerate(data, start=1):
            pct_master_tracker_sheet.cell(row=i, column=j).value = value

    # Guardar el nuevo archivo
    pct_master_tracker.save(f"PCT MASTER TRACKER {date}.xlsx")
    print(f"Datos combinados guardados en PCT MASTER TRACKER {date}.xlsx")
